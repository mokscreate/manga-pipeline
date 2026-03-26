"""
pipeline.py — 小说 → 推文文案 → 角色/场景 → 分镜 自动化工作流
================================================================
用法：
    python pipeline.py <小说文件.txt>

依赖：
    pip install -r requirements.txt

环境变量：
    ANTHROPIC_API_KEY=sk-ant-...   （Claude 步骤需要）
    DEEPSEEK_API_KEY=sk-...        （DeepSeek 步骤需要）
================================================================
"""

import os
import sys
import importlib
from pathlib import Path

from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import STEP_CONFIG, OUTPUT_EXCEL, SHEET_COLUMNS
from feishu import save_to_feishu, get_token, read_input_from_feishu, mark_input_done, read_prompts_from_feishu, init_prompts_to_feishu


# ── 客户端初始化 ──────────────────────────────────────────────

def make_clients() -> dict:
    """初始化 DeepSeek 客户端，缺少 Key 时给出明确提示。"""
    key = os.environ.get("DEEPSEEK_API_KEY")
    if not key:
        print("❌ 缺少环境变量 DEEPSEEK_API_KEY")
        sys.exit(1)
    return {
        "deepseek": OpenAI(
            api_key=key,
            base_url="https://api.deepseek.com",
        )
    }


# ── 统一调用入口 ──────────────────────────────────────────────

def call_model(clients: dict, step_key: str, prompt: str) -> str:
    """根据 STEP_CONFIG 自动选择 Claude 或 DeepSeek，流式输出并返回完整文本。"""
    cfg = STEP_CONFIG[step_key]
    provider = cfg["provider"]
    model = cfg["model"]
    max_tokens = cfg["max_tokens"]

    print(f"\n{'='*60}")
    print(f"  ▶ [{step_key}]  provider={provider}  model={model}")
    print(f"{'='*60}\n")

    full_text = ""

    client: OpenAI = clients["deepseek"]
    stream = client.chat.completions.create(
        model=model,
        max_tokens=max_tokens,
        messages=[{"role": "user", "content": prompt}],
        stream=True,
    )
    for chunk in stream:
        delta = chunk.choices[0].delta.content or ""
        print(delta, end="", flush=True)
        full_text += delta

    print(f"\n\n  ✓ {step_key} 完成（{len(full_text)} 字符）\n")
    return full_text


# ── Prompt 加载 ───────────────────────────────────────────────

# 运行时从飞书加载的prompt缓存，由main()填充
_feishu_prompts: dict = {}


def load_prompt(module_name: str) -> str:
    """优先从飞书prompt表读取，fallback到本地文件。"""
    if module_name in _feishu_prompts:
        return _feishu_prompts[module_name]
    mod = importlib.import_module(f"prompts.{module_name}")
    return mod.PROMPT_TEMPLATE


# ── 四个步骤 ──────────────────────────────────────────────────

def step1_novel_to_script(clients: dict, novel_text: str) -> str:
    prompt = load_prompt("step1_novel_to_script").format(novel_text=novel_text)
    return call_model(clients, "novel_to_script", prompt)


def step2_script_to_characters(clients: dict, script_text: str) -> str:
    prompt = load_prompt("step2_script_to_characters").format(script_text=script_text)
    return call_model(clients, "script_to_characters", prompt)


def step3_script_to_scenes(clients: dict, script_text: str) -> str:
    prompt = load_prompt("step3_script_to_scenes").format(script_text=script_text)
    return call_model(clients, "script_to_scenes", prompt)


def step4_scenes_to_storyboard(
    clients: dict,
    script_text: str,
    characters_text: str,
    scenes_text: str,
) -> str:
    prompt = load_prompt("step4_scenes_to_storyboard").format(
        script_text=script_text,
        characters_text=characters_text,
        scenes_text=scenes_text,
    )
    return call_model(clients, "scenes_to_storyboard", prompt)


# ── Excel 输出 ────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
HEADER_FONT  = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN   = Alignment(vertical="top", wrap_text=True)


def write_script_sheet(ws, script_text: str) -> int:
    """Step1 推文文案：段落格式，每段一行。"""
    for col_idx, name in enumerate(["序号", "段落内容"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    paragraphs = [p.strip() for p in script_text.splitlines() if p.strip()]
    for row_idx, para in enumerate(paragraphs, start=2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = CELL_ALIGN
        ws.cell(row=row_idx, column=2, value=para).alignment = CELL_ALIGN

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 60
    ws.row_dimensions[1].height = 22
    return len(paragraphs)


def parse_table(raw: str, columns: list) -> list:
    """解析 ||| 分隔的表格行，跳过空行和表头行。"""
    rows = []
    for line in raw.splitlines():
        line = line.strip()
        if not line or "|||" not in line:
            continue
        if any(col in line for col in columns[:3]):
            continue
        parts = [p.strip() for p in line.split("|||")]
        while len(parts) < len(columns):
            parts.append("")
        rows.append(dict(zip(columns, parts[: len(columns)])))
    return rows


def write_table_sheet(ws, columns: list, rows: list):
    """Step2/3/4 结构化表格。"""
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(name, ""))
            cell.alignment = CELL_ALIGN

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)

    ws.row_dimensions[1].height = 22


def save_excel(script_text, characters_text, scenes_text, storyboard_text, output_path):
    wb = Workbook()

    ws = wb.active
    ws.title = "推文文案"
    n = write_script_sheet(ws, script_text)
    print(f"  Sheet [推文文案]：{n} 段")

    for sheet_name, raw_text in [
        ("角色", characters_text),
        ("场景", scenes_text),
        ("分镜", storyboard_text),
    ]:
        ws = wb.create_sheet(sheet_name)
        cols = SHEET_COLUMNS[sheet_name]
        rows = parse_table(raw_text, cols)
        write_table_sheet(ws, cols, rows)
        print(f"  Sheet [{sheet_name}]：{len(rows)} 行")

    # 原始文本 Sheet，方便核查
    raw_ws = wb.create_sheet("原始输出")
    raw_ws.column_dimensions["A"].width = 15
    raw_ws.column_dimensions["B"].width = 80
    for r, (label, text) in enumerate([
        ("Step1 推文文案", script_text),
        ("Step2 角色",    characters_text),
        ("Step3 场景",    scenes_text),
        ("Step4 分镜",    storyboard_text),
    ], start=1):
        raw_ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        raw_ws.cell(row=r, column=2, value=text).alignment = CELL_ALIGN

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\n✅ Excel 已保存：{output_path}")


# ── 主入口 ────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    feishu_mode = "--feishu" in args

    if feishu_mode:
        print("🔗 正在从飞书读取剧本...")
        token = get_token()
        record_id, novel_text = read_input_from_feishu(token)
        print(f"📖 已读取飞书输入（{len(novel_text)} 字符）")
    elif args and not args[0].startswith("--"):
        novel_path = args[0]
        novel_text = Path(novel_path).read_text(encoding="utf-8")
        token = None
        record_id = None
        print(f"📖 已读取文件：{novel_path}（{len(novel_text)} 字符）")
    else:
        print("📋 请粘贴剧本内容，输入完成后按 Ctrl+Z 然后回车（Windows）或 Ctrl+D（Mac/Linux）：")
        novel_text = sys.stdin.read()
        token = None
        record_id = None
        print(f"📖 已读取输入（{len(novel_text)} 字符）")

    global _feishu_prompts
    try:
        _token = token or get_token()
        _feishu_prompts = read_prompts_from_feishu(_token)
        if not _feishu_prompts:
            print("📝 飞书prompt表为空，正在初始化...")
            import prompts.step1_novel_to_script as _p1
            import prompts.step2_script_to_characters as _p2
            import prompts.step3_script_to_scenes as _p3
            import prompts.step4_scenes_to_storyboard as _p4
            local_prompts = {
                "step1_novel_to_script":      _p1.PROMPT_TEMPLATE,
                "step2_script_to_characters": _p2.PROMPT_TEMPLATE,
                "step3_script_to_scenes":     _p3.PROMPT_TEMPLATE,
                "step4_scenes_to_storyboard": _p4.PROMPT_TEMPLATE,
            }
            init_prompts_to_feishu(_token, local_prompts)
            _feishu_prompts = local_prompts
        else:
            print(f"✅ 已从飞书加载 {len(_feishu_prompts)} 个prompt")
    except Exception as e:
        print(f"⚠ 飞书prompt加载失败，使用本地文件：{e}")
        _feishu_prompts = {}

    clients = make_clients()

    def update_status(msg):
        if record_id and token:
            mark_input_done(token, record_id, status=msg)

    def flush_to_feishu(sheet_name, fields, rows):
        if not token:
            return
        from feishu import get_token as _get_token, list_tables, ensure_table, clear_table, batch_insert
        t = _get_token()
        existing = list_tables(t)
        table_id = ensure_table(t, sheet_name, fields, existing)
        clear_table(t, table_id)
        batch_insert(t, table_id, rows)
        print(f"  ✓ 飞书[{sheet_name}]已写入 {len(rows)} 条")

    from feishu import parse_script_paragraphs, parse_table as feishu_parse_table

    print("\n[1/4] 正在生成推文文案...")
    update_status("⏳ Step1 生成推文文案中...")
    script_text = step1_novel_to_script(clients, novel_text)
    flush_to_feishu("推文文案", ["序号", "段落内容"], parse_script_paragraphs(script_text))

    print("\n[2/4] 正在提取角色设定...")
    update_status("⏳ Step2 提取角色设定中...")
    characters_text = step2_script_to_characters(clients, script_text)
    flush_to_feishu("角色", SHEET_COLUMNS["角色"], feishu_parse_table(characters_text, SHEET_COLUMNS["角色"]))

    print("\n[3/4] 正在提取场景设定...")
    update_status("⏳ Step3 提取场景设定中...")
    scenes_text = step3_script_to_scenes(clients, script_text)
    flush_to_feishu("场景", SHEET_COLUMNS["场景"], feishu_parse_table(scenes_text, SHEET_COLUMNS["场景"]))

    print("\n[4/4] 正在生成分镜脚本...")
    update_status("⏳ Step4 生成分镜脚本中...")
    storyboard_text = step4_scenes_to_storyboard(clients, script_text, characters_text, scenes_text)
    flush_to_feishu("分镜", SHEET_COLUMNS["分镜"], feishu_parse_table(storyboard_text, SHEET_COLUMNS["分镜"]))

    print("\n📊 正在生成 Excel...")
    save_excel(script_text, characters_text, scenes_text, storyboard_text, OUTPUT_EXCEL)

    if record_id and token:
        mark_input_done(token, record_id, status="已完成")
        print("  ✓ 输入记录已标记为「已完成」")


if __name__ == "__main__":
    main()
